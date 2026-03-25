"""
Export all Yaya Q&A entries from yaya_qa_full.md to an Excel spreadsheet.

Usage:
    python scripts/export_qa_excel.py
    python scripts/export_qa_excel.py --output my_file.xlsx
"""
import re
import argparse
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

QA_DOC = "data/memory/yaya_qa_full.md"
DEFAULT_OUT = "data/yaya_qa_export.xlsx"


def parse_qa_full(path):
    """Parse yaya_qa_full.md into a list of dicts."""
    with open(path, encoding="utf-8") as f:
        content = f.read()

    entries = []
    # Match: ### Q{n} — {domain}\n**{question}**\n\n{answer}\n\n**Rating: {r}/10**
    pattern = re.compile(
        r"### Q(\d+)\s+[—–-]+\s+(.+?)\n"   # Q number and domain
        r"\*\*(.+?)\*\*\n\n"                  # question (bold)
        r"(.*?)\n\n"                           # answer
        r"\*\*Rating:\s*([\d.]+)(?:/10)?",    # rating
        re.DOTALL,
    )

    for m in pattern.finditer(content):
        q_num = int(m.group(1))
        domain = m.group(2).strip()
        question = m.group(3).strip()
        answer = m.group(4).strip()
        rating_str = m.group(5).strip()
        try:
            rating = float(rating_str)
        except ValueError:
            rating = None
        entries.append({
            "q_num": q_num,
            "domain": domain,
            "question": question,
            "answer": answer,
            "rating": rating,
        })

    return sorted(entries, key=lambda x: x["q_num"])


def domain_color(domain):
    """Return a hex fill color for each domain."""
    palette = {
        "Physics": "DDEEFF",
        "Biology": "DDFFDD",
        "Chemistry": "FFEEDD",
        "Mathematics": "F0DDFF",
        "Computer Science": "FFDDEE",
        "Philosophy": "FFFACC",
        "Psychology": "FFE4CC",
        "Neuroscience": "CCF5FF",
        "History": "F5DDCC",
        "Economics": "DDFFF0",
        "Linguistics": "EEF5DD",
        "Medicine": "FFD9D9",
        "Ecology": "D9FFE8",
        "Geography": "D9EEFF",
        "Astronomy": "E8D9FF",
    }
    return palette.get(domain, "F5F5F5")


def build_excel(entries, output_path):
    wb = openpyxl.Workbook()

    # ── Sheet 1: Full Q&A ────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Q&A Full"

    headers = ["#", "Domain", "Question", "Answer", "Rating"]
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="2F4F8F")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row_idx, e in enumerate(entries, 2):
        color = domain_color(e["domain"])
        row_fill = PatternFill("solid", fgColor=color)

        cells = [
            ws.cell(row=row_idx, column=1, value=e["q_num"]),
            ws.cell(row=row_idx, column=2, value=e["domain"]),
            ws.cell(row=row_idx, column=3, value=e["question"]),
            ws.cell(row=row_idx, column=4, value=e["answer"]),
            ws.cell(row=row_idx, column=5, value=e["rating"]),
        ]
        for c in cells:
            c.fill = row_fill
            c.border = border
            c.alignment = Alignment(vertical="top", wrap_text=True)

        cells[0].alignment = Alignment(horizontal="center", vertical="top")
        cells[4].alignment = Alignment(horizontal="center", vertical="top")

    # Column widths
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 42
    ws.column_dimensions["D"].width = 80
    ws.column_dimensions["E"].width = 8

    # ── Sheet 2: Summary by Domain ───────────────────────────────────────────
    ws2 = wb.create_sheet("By Domain")
    from collections import defaultdict
    by_domain = defaultdict(list)
    for e in entries:
        by_domain[e["domain"]].append(e)

    ws2.append(["Domain", "Count", "Avg Rating", "Q Numbers"])
    for cell in ws2[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2F4F8F")
        cell.alignment = Alignment(horizontal="center")

    for domain in sorted(by_domain):
        items = by_domain[domain]
        ratings = [e["rating"] for e in items if e["rating"] is not None]
        avg = round(sum(ratings) / len(ratings), 2) if ratings else ""
        q_nums = ", ".join(str(e["q_num"]) for e in items)
        row = ws2.append([domain, len(items), avg, q_nums])
        color = domain_color(domain)
        fill = PatternFill("solid", fgColor=color)
        for cell in ws2[ws2.max_row]:
            cell.fill = fill
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 8
    ws2.column_dimensions["C"].width = 12
    ws2.column_dimensions["D"].width = 80
    ws2.freeze_panes = "A2"

    wb.save(output_path)
    return len(entries)


def main():
    parser = argparse.ArgumentParser(description="Export Yaya Q&A to Excel")
    parser.add_argument("--output", default=DEFAULT_OUT, help="Output .xlsx path")
    parser.add_argument("--source", default=QA_DOC, help="Source markdown file")
    args = parser.parse_args()

    print(f"Reading: {args.source}")
    entries = parse_qa_full(args.source)
    print(f"Parsed {len(entries)} Q&A entries")

    build_excel(entries, args.output)
    print(f"Saved: {args.output}")
    print(f"  Sheet 1 'Q&A Full':    {len(entries)} rows")
    print(f"  Sheet 2 'By Domain':   {len(set(e['domain'] for e in entries))} domains")


if __name__ == "__main__":
    main()
